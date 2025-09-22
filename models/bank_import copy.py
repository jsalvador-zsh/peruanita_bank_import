import base64
import io
from datetime import datetime
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import logging

_logger = logging.getLogger(__name__)


class BankImport(models.Model):
    _name = 'bank.import'
    _description = 'Importación de Operaciones Bancarias'
    _order = 'create_date desc'

    name = fields.Char('Nombre', required=True, default=lambda: _('Nueva Importación'))
    file_data = fields.Binary('Archivo', required=True)
    file_name = fields.Char('Nombre del Archivo')
    file_type = fields.Selection([
        ('txt', 'Archivo TXT'),
        ('excel', 'Archivo Excel')
    ], string='Tipo de Archivo', compute='_compute_file_type', store=True)
    
    bank_type = fields.Selection([
        ('bcp', 'Banco de Crédito del Perú'),
        ('nacion', 'Banco de la Nación'),
        ('other', 'Otro Banco')
    ], string='Banco', required=True, default='bcp')
    
    state = fields.Selection([
        ('draft', 'Borrador'),
        ('processed', 'Procesado'),
        ('matched', 'Emparejado')
    ], string='Estado', default='draft')
    
    import_date = fields.Datetime('Fecha de Importación', default=fields.Datetime.now)
    line_ids = fields.One2many('bank.import.line', 'import_id', string='Líneas de Importación')
    matched_payment_ids = fields.One2many('bank.import.match', 'import_id', string='Pagos Emparejados')
    
    total_operations = fields.Integer('Total Operaciones', compute='_compute_totals')
    matched_operations = fields.Integer('Operaciones Emparejadas', compute='_compute_totals')
    unmatched_operations = fields.Integer('Operaciones Sin Emparejar', compute='_compute_totals')

    @api.depends('file_name')
    def _compute_file_type(self):
        for record in self:
            if record.file_name:
                if record.file_name.lower().endswith('.txt'):
                    record.file_type = 'txt'
                elif record.file_name.lower().endswith(('.xls', '.xlsx')):
                    record.file_type = 'excel'
                else:
                    record.file_type = 'txt'
            else:
                record.file_type = 'txt'

    @api.depends('line_ids', 'matched_payment_ids')
    def _compute_totals(self):
        for record in self:
            record.total_operations = len(record.line_ids)
            record.matched_operations = len(record.matched_payment_ids)
            record.unmatched_operations = record.total_operations - record.matched_operations

    def action_process_file(self):
        """Procesar el archivo importado"""
        if not self.file_data:
            raise UserError(_('Debe cargar un archivo antes de procesarlo.'))
        
        _logger.info(f"=== INICIANDO PROCESAMIENTO DE ARCHIVO ===")
        _logger.info(f"Tipo de archivo: {self.file_type}")
        _logger.info(f"Nombre del archivo: {self.file_name}")
        _logger.info(f"Banco: {self.bank_type}")
        
        self.line_ids.unlink()  # Limpiar líneas anteriores
        
        try:
            if self.file_type == 'txt':
                _logger.info("Procesando archivo TXT...")
                self._process_txt_file()
            elif self.file_type == 'excel':
                _logger.info("Procesando archivo Excel...")
                self._process_excel_file()
            
            lines_count = len(self.line_ids)
            _logger.info(f"=== PROCESAMIENTO COMPLETADO: {lines_count} líneas creadas ===")
            
            if lines_count == 0:
                raise UserError(_('No se pudieron extraer datos del archivo. Verifique que el archivo tenga el formato correcto y contenga datos.'))
            
            self.state = 'processed'
            self.name = f"Importación {self.bank_type.upper()} - {fields.Datetime.now().strftime('%d/%m/%Y %H:%M')}"
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Éxito'),
                    'message': _('Se procesaron %d operaciones del archivo.') % lines_count,
                    'type': 'success',
                    'sticky': False,
                }
            }
            
        except Exception as e:
            _logger.error(f"=== ERROR EN PROCESAMIENTO: {str(e)} ===")
            raise

    def _process_txt_file(self):
        """Procesar archivo TXT del banco"""
        try:
            file_content = base64.b64decode(self.file_data).decode('utf-8')
            lines = file_content.split('\n')
            
            _logger.info(f"Procesando archivo TXT para banco: {self.bank_type}")
            
            # Buscar líneas de transacciones (que empiecen con fecha)
            transaction_lines = []
            for line in lines:
                if line.strip() and line.startswith('"') and self._is_transaction_line(line):
                    transaction_lines.append(line)
            
            _logger.info(f"Encontradas {len(transaction_lines)} líneas de transacciones")
            
            for line in transaction_lines:
                self._parse_txt_transaction(line)
                
        except Exception as e:
            _logger.error(f"Error procesando archivo TXT: {str(e)}")
            raise UserError(_('Error al procesar el archivo TXT: %s') % str(e))

    def _is_transaction_line(self, line):
        """Verificar si una línea es una transacción válida"""
        try:
            # Dividir por punto y coma y verificar formato de fecha
            fields = [field.strip('"') for field in line.split(';')]
            if len(fields) >= 6:
                date_str = fields[0]
                # Verificar formato de fecha DD/MM/YYYY
                datetime.strptime(date_str, '%d/%m/%Y')
                return True
        except:
            pass
        return False

    def _parse_txt_transaction(self, line):
        """Parsear una línea de transacción del TXT"""
        try:
            fields = [field.strip('"').strip() for field in line.split(';')]
            
            if len(fields) >= 6:
                date_str = fields[0]
                description = fields[2]
                amount_str = fields[3].replace(',', '')
                operation_number = fields[5]
                
                # Convertir fecha
                transaction_date = datetime.strptime(date_str, '%d/%m/%Y').date()
                
                # Convertir monto
                try:
                    amount = float(amount_str)
                except ValueError:
                    amount = 0.0
                
                # Para BCP: Solo tomar los últimos 6 dígitos del número de operación
                if self.bank_type == 'bcp' and operation_number and len(operation_number) >= 6:
                    operation_number = operation_number[-6:]
                    _logger.info(f"BCP: Número de operación ajustado a últimos 6 dígitos: {operation_number}")
                
                # Crear línea de importación
                self.env['bank.import.line'].create({
                    'import_id': self.id,
                    'transaction_date': transaction_date,
                    'description': description,
                    'amount': amount,
                    'operation_number': operation_number,
                    'original_line': line
                })
                
        except Exception as e:
            _logger.error(f"Error parseando línea TXT: {str(e)}")

    def _process_excel_file(self):
        """Procesar archivo Excel del Banco de la Nación"""
        try:
            file_content = base64.b64decode(self.file_data)
            _logger.info(f"Procesando archivo Excel de {len(file_content)} bytes")
            
            # Verificar qué librerías están disponibles
            openpyxl_available = False
            xlrd_available = False
            
            try:
                import openpyxl
                openpyxl_available = True
                _logger.info(f"Librería openpyxl {openpyxl.__version__} disponible")
            except ImportError as e:
                _logger.warning(f"Librería openpyxl no disponible: {e}")
            
            try:
                import xlrd
                xlrd_available = True
                _logger.info(f"Librería xlrd {xlrd.__version__} disponible")
            except ImportError as e:
                _logger.warning(f"Librería xlrd no disponible: {e}")
            
            if not openpyxl_available and not xlrd_available:
                raise UserError(_('No se encontraron librerías para procesar archivos Excel.'))
            
            success = False
            last_error = None
            
            # Intentar con openpyxl primero (para .xlsx)
            if openpyxl_available:
                try:
                    import openpyxl
                    _logger.info("Intentando procesar con openpyxl...")
                    workbook = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
                    sheet = workbook.active
                    _logger.info(f"Archivo Excel procesado con openpyxl. Hoja activa: {sheet.title}, Filas: {sheet.max_row}, Columnas: {sheet.max_column}")
                    
                    if sheet.max_row < 2:
                        raise UserError(_('El archivo Excel está vacío o no contiene datos suficientes.'))
                    
                    self._parse_excel_openpyxl(sheet)
                    success = True
                    _logger.info("Procesamiento con openpyxl exitoso")
                    
                except Exception as e:
                    last_error = str(e)
                    _logger.warning(f"Error con openpyxl: {last_error}")
                    if "not a zip file" in last_error.lower() or "invalid" in last_error.lower():
                        # Es probable que sea un archivo .xls, intentar con xlrd
                        pass
                    elif not xlrd_available:
                        raise UserError(_('Error procesando archivo Excel: %s\n\nVerifique que el archivo no esté corrupto.') % last_error)
            
            # Intentar con xlrd para archivos .xls
            if not success and xlrd_available:
                try:
                    import xlrd
                    _logger.info("Intentando procesar con xlrd...")
                    workbook = xlrd.open_workbook(file_contents=file_content)
                    sheet = workbook.sheet_by_index(0)
                    _logger.info(f"Archivo Excel procesado con xlrd. Hoja: {sheet.name}, Filas: {sheet.nrows}, Columnas: {sheet.ncols}")
                    
                    if sheet.nrows < 2:
                        raise UserError(_('El archivo Excel está vacío o no contiene datos suficientes.'))
                    
                    self._parse_excel_xlrd(sheet)
                    success = True
                    _logger.info("Procesamiento con xlrd exitoso")
                    
                except Exception as e:
                    last_error = str(e)
                    _logger.error(f"Error con xlrd: {last_error}")
            
            if not success:
                error_msg = _('No se pudo procesar el archivo Excel.')
                if last_error:
                    error_msg += f'\n\nÚltimo error: {last_error}'
                error_msg += _('\n\nVerifique que:\n- El archivo no esté corrupto\n- El archivo tenga extensión .xlsx o .xls\n- El archivo contenga datos válidos')
                raise UserError(error_msg)
                
        except UserError:
            raise  # Re-raise UserError tal como está
        except Exception as e:
            _logger.error(f"Error inesperado procesando archivo Excel: {str(e)}")
            raise UserError(_('Error inesperado al procesar el archivo Excel: %s') % str(e))

    def _parse_excel_openpyxl(self, sheet):
        """Parsear Excel con openpyxl"""
        _logger.info("Iniciando parseado con openpyxl...")
        
        # Buscar header row
        header_row = None
        for row_num in range(1, min(10, sheet.max_row + 1)):
            row = sheet[row_num]
            row_values = [str(cell.value).lower() if cell.value else '' for cell in row]
            _logger.info(f"Fila {row_num}: {row_values}")
            
            if any('fecha' in value for value in row_values):
                header_row = row_num
                _logger.info(f"Header encontrado en fila {row_num}")
                break
        
        if not header_row:
            raise UserError(_('No se encontró la fila de encabezados en el archivo Excel. Verifique que haya una columna con "fecha".'))
        
        # Obtener headers
        headers = [str(cell.value).lower().strip() if cell.value else '' for cell in sheet[header_row]]
        _logger.info(f"Headers encontrados: {headers}")
        
        # Mapear columnas
        col_mapping = self._get_excel_column_mapping(headers)
        _logger.info(f"Mapeo de columnas: {col_mapping}")
        
        if not col_mapping:
            raise UserError(_('No se pudieron mapear las columnas del Excel. Verifique que contenga las columnas necesarias (fecha, descripción, monto, operación).'))
        
        # Procesar datos
        lines_created = 0
        for row_num in range(header_row + 1, sheet.max_row + 1):
            row = sheet[row_num]
            if not any(cell.value for cell in row):
                continue
            
            try:
                self._create_excel_import_line(row, col_mapping, 'openpyxl')
                lines_created += 1
            except Exception as e:
                _logger.warning(f"Error procesando fila {row_num}: {str(e)}")
                
        _logger.info(f"Se crearon {lines_created} líneas desde Excel")
        
        if lines_created == 0:
            raise UserError(_('No se pudieron extraer datos del archivo Excel. Verifique el formato de los datos.'))

    def _parse_excel_xlrd(self, sheet):
        """Parsear Excel con xlrd"""
        _logger.info("Iniciando parseado con xlrd...")
        
        # Buscar header row
        header_row = None
        for row_num in range(min(10, sheet.nrows)):
            row = sheet.row_values(row_num)
            row_values = [str(cell).lower() for cell in row]
            _logger.info(f"Fila {row_num}: {row_values}")
            
            if any('fecha' in value for value in row_values):
                header_row = row_num
                _logger.info(f"Header encontrado en fila {row_num}")
                break
        
        if header_row is None:
            raise UserError(_('No se encontró la fila de encabezados en el archivo Excel. Verifique que haya una columna con "fecha".'))
        
        # Obtener headers
        headers = [str(cell).lower().strip() for cell in sheet.row_values(header_row)]
        _logger.info(f"Headers encontrados: {headers}")
        
        # Mapear columnas
        col_mapping = self._get_excel_column_mapping(headers)
        _logger.info(f"Mapeo de columnas: {col_mapping}")
        
        if not col_mapping:
            raise UserError(_('No se pudieron mapear las columnas del Excel. Verifique que contenga las columnas necesarias (fecha, descripción, monto, operación).'))
        
        # Procesar datos
        lines_created = 0
        for row_num in range(header_row + 1, sheet.nrows):
            row = sheet.row_values(row_num)
            if not any(cell for cell in row):
                continue
            
            try:
                self._create_excel_import_line(row, col_mapping, 'xlrd')
                lines_created += 1
            except Exception as e:
                _logger.warning(f"Error procesando fila {row_num}: {str(e)}")
                
        _logger.info(f"Se crearon {lines_created} líneas desde Excel")
        
        if lines_created == 0:
            raise UserError(_('No se pudieron extraer datos del archivo Excel. Verifique el formato de los datos.'))

    def _get_excel_column_mapping(self, headers):
        """Mapear columnas del Excel"""
        mapping = {}
        
        _logger.info(f"Mapeando headers: {headers}")
        
        for i, header in enumerate(headers):
            header_lower = header.lower().strip()
            
            # Mapear fecha
            if any(word in header_lower for word in ['fecha', 'date', 'dia']):
                mapping['date'] = i
                _logger.info(f"Columna fecha mapeada: {i} ({header})")
            
            # Mapear descripción (Trans. en este caso)
            elif any(word in header_lower for word in ['descripcion', 'concepto', 'detalle', 'description', 'memo', 'glosa', 'trans']):
                mapping['description'] = i
                _logger.info(f"Columna descripción mapeada: {i} ({header})")
            
            # Mapear cargo (débitos)
            elif any(word in header_lower for word in ['cargo', 'debe', 'debito']):
                mapping['cargo'] = i
                _logger.info(f"Columna cargo mapeada: {i} ({header})")
            
            # Mapear abono (créditos)
            elif any(word in header_lower for word in ['abono', 'haber', 'credito']):
                mapping['abono'] = i
                _logger.info(f"Columna abono mapeada: {i} ({header})")
            
            # Mapear número de operación/documento
            elif any(word in header_lower for word in ['documento', 'nro', 'numero', 'referencia', 'reference', 'operation']):
                mapping['operation'] = i
                _logger.info(f"Columna operación mapeada: {i} ({header})")
        
        _logger.info(f"Mapeo final: {mapping}")
        
        # Verificar que tenemos al menos fecha
        if 'date' not in mapping:
            _logger.warning("No se encontró columna de fecha")
        
        return mapping

    def _create_excel_import_line(self, row, col_mapping, parser_type):
        """Crear línea de importación desde Excel"""
        try:
            # Extraer datos según el parser
            if parser_type == 'openpyxl':
                get_value = lambda i: row[i].value if i < len(row) and row[i].value is not None else ''
            else:  # xlrd
                get_value = lambda i: row[i] if i < len(row) else ''
            
            _logger.info(f"Procesando fila con parser {parser_type}")
            
            # Fecha
            transaction_date = None
            if 'date' in col_mapping:
                date_value = get_value(col_mapping['date'])
                _logger.info(f"Valor fecha crudo: {date_value} (tipo: {type(date_value)})")
                
                if date_value:
                    if isinstance(date_value, datetime):
                        transaction_date = date_value.date()
                    elif isinstance(date_value, str):
                        # Intentar varios formatos de fecha
                        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y']:
                            try:
                                transaction_date = datetime.strptime(str(date_value).strip(), fmt).date()
                                break
                            except:
                                continue
                    # Para xlrd, a veces las fechas vienen como números
                    elif isinstance(date_value, (int, float)) and parser_type == 'xlrd':
                        try:
                            import xlrd
                            date_tuple = xlrd.xldate_as_tuple(date_value, 0)
                            transaction_date = datetime(*date_tuple[:3]).date()
                        except:
                            pass
                            
                _logger.info(f"Fecha procesada: {transaction_date}")
            
            # Descripción
            description = ''
            if 'description' in col_mapping:
                desc_value = get_value(col_mapping['description'])
                description = str(desc_value) if desc_value else ''
                _logger.info(f"Descripción: {description}")
            
            # Monto
            amount = 0.0
            if 'amount' in col_mapping:
                amount_value = get_value(col_mapping['amount'])
                _logger.info(f"Valor monto crudo: {amount_value} (tipo: {type(amount_value)})")
                
                if amount_value:
                    try:
                        # Limpiar formato de monto
                        amount_str = str(amount_value).replace(',', '').replace(' ', '').strip()
                        amount = float(amount_str)
                    except:
                        amount = 0.0

                _logger.info(f"Monto procesado: {amount}")

            # Manejar columnas separadas de cargo y abono
            if 'cargo' in col_mapping:
                cargo_value = get_value(col_mapping['cargo'])
                if cargo_value:
                    try:
                        cargo_str = str(cargo_value).replace(',', '').replace(' ', '').strip()
                        cargo_amount = float(cargo_str)
                        if cargo_amount > 0:
                            amount = -cargo_amount  # Cargo es negativo
                    except:
                        pass

            if 'abono' in col_mapping:
                abono_value = get_value(col_mapping['abono'])
                if abono_value:
                    try:
                        abono_str = str(abono_value).replace(',', '').replace(' ', '').strip()
                        abono_amount = float(abono_str)
                        if abono_amount > 0:
                            amount = abono_amount  # Abono es positivo
                    except:
                        pass

            # Número de operación
            operation_number = ''
            if 'operation' in col_mapping:
                op_value = get_value(col_mapping['operation'])
                operation_number = str(op_value) if op_value else ''
                _logger.info(f"Número de operación: {operation_number}")

            # Crear línea si tenemos datos mínimos
            if transaction_date or amount != 0 or operation_number:
                line_vals = {
                    'import_id': self.id,
                    'transaction_date': transaction_date or fields.Date.today(),
                    'description': description,
                    'amount': amount,
                    'operation_number': operation_number,
                    'original_line': f"Excel row: {str(row)}"
                }

                _logger.info(f"Creando línea con valores: {line_vals}")
                new_line = self.env['bank.import.line'].create(line_vals)
                _logger.info(f"Línea creada exitosamente: ID {new_line.id}")
            else:
                _logger.warning("Fila descartada: no contiene datos suficientes")

        except Exception as e:
            _logger.error(f"Error creando línea Excel: {str(e)}")
            _logger.error(f"Datos de la fila: {row}")
            _logger.error(f"Mapeo de columnas: {col_mapping}")
            raise

    def action_debug_excel(self):
        """Método de debug para analizar archivo Excel"""
        if not self.file_data:
            raise UserError(_('Debe cargar un archivo primero.'))
        
        try:
            file_content = base64.b64decode(self.file_data)
            
            # Información básica del archivo
            debug_info = []
            debug_info.append(f"Tamaño del archivo: {len(file_content)} bytes")
            debug_info.append(f"Primeros 50 bytes (hex): {file_content[:50].hex()}")
            debug_info.append(f"Primeros 50 bytes (text): {file_content[:50]}")
            
            # Intentar con openpyxl
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
                sheet = workbook.active
                debug_info.append(f"✅ openpyxl: Hoja '{sheet.title}', {sheet.max_row} filas, {sheet.max_column} columnas")
                
                # Mostrar primeras 5 filas
                for i in range(1, min(6, sheet.max_row + 1)):
                    row_data = [str(cell.value) for cell in sheet[i]]
                    debug_info.append(f"Fila {i}: {row_data}")
                    
            except Exception as e:
                debug_info.append(f"❌ openpyxl error: {str(e)}")
            
            # Intentar con xlrd
            try:
                import xlrd
                workbook = xlrd.open_workbook(file_contents=file_content)
                sheet = workbook.sheet_by_index(0)
                debug_info.append(f"✅ xlrd: Hoja '{sheet.name}', {sheet.nrows} filas, {sheet.ncols} columnas")
                
                # Mostrar primeras 5 filas
                for i in range(min(5, sheet.nrows)):
                    row_data = [str(cell) for cell in sheet.row_values(i)]
                    debug_info.append(f"Fila {i}: {row_data}")
                    
            except Exception as e:
                debug_info.append(f"❌ xlrd error: {str(e)}")
                
            message = '\n'.join(debug_info)
            _logger.info(f"DEBUG EXCEL:\n{message}")
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Debug Excel'),
                    'message': message,
                    'type': 'info',
                    'sticky': True,
                }
            }

        except Exception as e:
            _logger.error(f"Error en debug: {str(e)}")
            raise UserError(_('Error analizando archivo: %s') % str(e))

    def action_match_payments(self):
        """Buscar pagos que coincidan con las operaciones importadas"""
        if not self.line_ids:
            raise UserError(_('Debe procesar el archivo primero.'))

        _logger.info(f"Iniciando búsqueda de matches para importación {self.id}")

        # Limpiar matches anteriores
        self.matched_payment_ids.unlink()

        total_matches = 0
        for line in self.line_ids:
            try:
                matches_count = self._find_matching_payments(line)
                total_matches += matches_count
            except Exception as e:
                _logger.error(f"Error procesando línea {line.id}: {str(e)}")
                raise UserError(_('Error procesando línea de operación %s: %s') % (line.operation_number, str(e)))

        self.state = 'matched'
        _logger.info(f"Búsqueda completada. Total matches encontrados: {total_matches}")

        # Retornar True para que se refresque la vista automáticamente
        return True

    def _find_matching_payments(self, import_line):
        """Buscar pagos que coincidan con una línea de importación"""
        _logger.info(f"Buscando matches para línea: {import_line.id}, operación: {import_line.operation_number}, monto: {import_line.amount}")
        
        # Buscar en account.payment
        domain = [('state', 'in', ['posted', 'sent', 'in_process'])]
        
        # Buscar por monto exacto (valor absoluto)
        amount_matches = self.env['account.payment'].search(domain + [
            ('amount', '=', abs(import_line.amount))
        ])
        
        _logger.info(f"Encontrados {len(amount_matches)} pagos con monto {abs(import_line.amount)}")
        
        # También buscar con una tolerancia mínima para errores de redondeo
        if not amount_matches:
            tolerance = 0.01  # 1 centavo de tolerancia
            amount_matches = self.env['account.payment'].search(domain + [
                ('amount', '>=', abs(import_line.amount) - tolerance),
                ('amount', '<=', abs(import_line.amount) + tolerance)
            ])
            _logger.info(f"Con tolerancia encontrados {len(amount_matches)} pagos")
        
        matches_created = 0
        for payment in amount_matches:
            _logger.info(f"Verificando pago {payment.id} - {payment.name}, memo: {payment.memo}")
            
            # Verificar si el número de operación coincide en algún campo
            operation_match = self._check_operation_match(payment, import_line.operation_number)
            _logger.info(f"Coincidencia de operación: {operation_match}")
            
            # Crear match si coincide operación O si no hay número de operación
            if operation_match or not import_line.operation_number:
                # Verificar si ya existe este match para evitar duplicados
                existing_match = self.env['bank.import.match'].search([
                    ('import_id', '=', self.id),
                    ('payment_id', '=', payment.id),
                    ('import_line_id', '=', import_line.id)
                ])
                
                if not existing_match:
                    try:
                        match_vals = {
                            'import_id': self.id,
                            'import_line_id': import_line.id,
                            'payment_id': payment.id,
                            'match_type': 'exact' if operation_match else 'partial'
                        }
                        _logger.info(f"Creando match con valores: {match_vals}")
                        
                        new_match = self.env['bank.import.match'].create(match_vals)
                        matches_created += 1
                        _logger.info(f"Match creado exitosamente: {new_match.id}")
                    except Exception as e:
                        _logger.error(f"Error creando match: {str(e)}")
                        raise UserError(_('Error creando coincidencia: %s') % str(e))
        
        # Si no encontramos matches por monto, buscar solo por número de operación
        if matches_created == 0 and import_line.operation_number:
            _logger.info("Buscando matches solo por número de operación")
            all_payments = self.env['account.payment'].search(domain)
            for payment in all_payments:
                if self._check_operation_match(payment, import_line.operation_number):
                    existing_match = self.env['bank.import.match'].search([
                        ('import_id', '=', self.id),
                        ('payment_id', '=', payment.id),
                        ('import_line_id', '=', import_line.id)
                    ])
                    
                    if not existing_match:
                        try:
                            match_vals = {
                                'import_id': self.id,
                                'import_line_id': import_line.id,
                                'payment_id': payment.id,
                                'match_type': 'partial'
                            }
                            new_match = self.env['bank.import.match'].create(match_vals)
                            matches_created += 1
                            _logger.info(f"Match parcial creado: {new_match.id}")
                        except Exception as e:
                            _logger.error(f"Error creando match parcial: {str(e)}")
        
        _logger.info(f"Total de matches creados para línea {import_line.id}: {matches_created}")
        return matches_created

    def _check_operation_match(self, payment, operation_number):
        """Verificar si el número de operación coincide"""
        if not operation_number:
            return True  # Si no hay número de operación, considerar match por monto
        
        # Limpiar el número de operación (remover ceros a la izquierda y espacios)
        clean_operation = str(operation_number).strip().lstrip('0') or '0'
        
        # Campos donde puede estar el número de operación en account.payment
        fields_to_check = [
            payment.name,  # Número/referencia del pago
            payment.memo,  # Campo memo (donde está el número de operación)
        ]
        
        # Verificar también en el campo communication si existe
        if hasattr(payment, 'communication') and payment.communication:
            fields_to_check.append(payment.communication)
            
        # Verificar en payment_reference si existe
        if hasattr(payment, 'payment_reference') and payment.payment_reference:
            fields_to_check.append(payment.payment_reference)
        
        for field_value in fields_to_check:
            if field_value:
                field_str = str(field_value).strip()
                
                # Buscar coincidencia exacta
                if operation_number in field_str:
                    return True
                    
                # Buscar coincidencia sin ceros a la izquierda
                if clean_operation != '0' and clean_operation in field_str:
                    return True
                    
                # Buscar si el campo contiene el número sin ceros a la izquierda
                clean_field = field_str.lstrip('0') or '0'
                if clean_operation == clean_field:
                    return True
                
                # Para BCP: También buscar por los últimos 6 dígitos del campo del pago
                if len(field_str) >= 6:
                    last_6_digits = field_str[-6:]
                    if operation_number == last_6_digits or clean_operation == last_6_digits.lstrip('0'):
                        _logger.info(f"Coincidencia BCP por últimos 6 dígitos: {operation_number} == {last_6_digits}")
                        return True

        # Si no hay coincidencia exacta, permitir match solo por monto
        return False


class BankImportLine(models.Model):
    _name = 'bank.import.line'
    _description = 'Línea de Importación Bancaria'

    import_id = fields.Many2one('bank.import', string='Importación', required=True, ondelete='cascade')
    transaction_date = fields.Date('Fecha Transacción', required=True)
    description = fields.Char('Descripción')
    amount = fields.Float('Monto', digits=(16, 2))
    operation_number = fields.Char('Número de Operación')
    original_line = fields.Text('Línea Original')
    is_matched = fields.Boolean('Emparejado', compute='_compute_is_matched')

    @api.depends('import_id.matched_payment_ids')
    def _compute_is_matched(self):
        for record in self:
            record.is_matched = bool(record.import_id.matched_payment_ids.filtered(
                lambda m: m.import_line_id.id == record.id
            ))


class BankImportMatch(models.Model):
    _name = 'bank.import.match'
    _description = 'Emparejamiento de Pagos'

    import_id = fields.Many2one('bank.import', string='Importación', required=True, ondelete='cascade')
    import_line_id = fields.Many2one('bank.import.line', string='Línea de Importación', required=True, ondelete='cascade')
    payment_id = fields.Many2one('account.payment', string='Pago', required=True)
    match_type = fields.Selection([
        ('exact', 'Exacto'),
        ('partial', 'Parcial')
    ], string='Tipo de Coincidencia', default='exact')

    # Campos relacionados para mostrar información
    transaction_date = fields.Date(related='import_line_id.transaction_date', store=True)
    operation_number = fields.Char(related='import_line_id.operation_number', store=True)
    amount = fields.Float(related='import_line_id.amount', store=True)
    payment_amount = fields.Monetary(related='payment_id.amount', store=True, currency_field='currency_id')
    currency_id = fields.Many2one(related='payment_id.currency_id', store=True)
    payment_reference = fields.Char(related='payment_id.name', store=True, string='Referencia de Pago')
    payment_memo = fields.Char(related='payment_id.memo', store=True, string='Memo del Pago')
    partner_name = fields.Char(related='payment_id.partner_id.name', store=True)

    @api.model
    def create(self, vals):
        """Validar campos obligatorios antes de crear"""
        if not vals.get('import_line_id'):
            raise UserError(_('El campo "Línea de Importación" es obligatorio.'))
        if not vals.get('payment_id'):
            raise UserError(_('El campo "Pago" es obligatorio.'))
        if not vals.get('import_id'):
            raise UserError(_('El campo "Importación" es obligatorio.'))

        return super().create(vals)